import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def remove_empty_folders(root_path):
    for dirpath, dirnames, filenames in os.walk(root_path, topdown=False):
        for dirname in dirnames:
            folder_to_check = os.path.join(dirpath, dirname)
            if not os.listdir(folder_to_check):
                os.rmdir(folder_to_check)

df = pd.read_csv('weather_data.csv')
df = df.dropna()

cities = list(set(df['Location']))
cities.sort()

months = {
    1: 'january',
    2: 'february',
    3: 'march',
    4: 'april',
    5: 'may',
    6: 'june',
    7: 'july',
    8: 'august',
    9: 'september',
    10: 'october',
    11: 'november',
    12: 'december'
}

df['Date_Time'] = pd.to_datetime(df['Date'] + ' ' + df['Time'])

wb = Workbook()

default_sheet = "Sheet"
if default_sheet in wb.sheetnames:
    sheet_to_remove = wb[default_sheet]
    wb.remove(sheet_to_remove)

wb.create_sheet('All month long')
data_sheet = wb['All month long']

for city in cities:
    Path(f'Analyzed/{city}').mkdir(exist_ok=True)

    filtered_df = df[df['Location'] == city]
    filtered_df = filtered_df.sort_values(by='Date')
    
    for i in range(1, 13):
        Path(f'Analyzed/{city}/{months[i].capitalize()}').mkdir(exist_ok=True)
        
        data_sheet['A1'].value = 'Average temperature for the month (°C)'
        data_sheet['A1'].font = Font(bold=True)
        data_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_sheet['A2'].value = 'Average humidity for the month (%)'
        data_sheet['A2'].font = Font(bold=True)
        data_sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_sheet['A3'].value = 'Average monthly precipitation (mm)'
        data_sheet['A3'].font = Font(bold=True)
        data_sheet['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_sheet['A4'].value = 'Average wind speed for the month (km/h)'
        data_sheet['A4'].font = Font(bold=True)
        data_sheet['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        filtered_df = df[(df['Location'] == city) & (df['Date_Time'].dt.month == i)]
        
        temperature_avg = filtered_df['Temperature_C'].mean()
        data_sheet['B1'] = temperature_avg
        data_sheet['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        humididty_avg = filtered_df['Humidity_pct'].mean()
        data_sheet['B2'] = humididty_avg
        data_sheet['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        precipitation_avg = filtered_df['Precipitation_mm'].mean()
        data_sheet['B3'] = precipitation_avg
        data_sheet['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        wind_speed_avg = filtered_df['Wind_Speed_kmh'].mean()
        data_sheet['B4'] = wind_speed_avg
        data_sheet['B4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        unique_dates = list(set(filtered_df['Date']))
        unique_dates.sort()
        
        j = 1
        
        for date in unique_dates:
            data_sheet['A6'].value = 'Date'
            data_sheet['A6'].font = Font(bold=True)
            data_sheet['A6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['B6'].value = 'Average temperature (°C)'
            data_sheet['B6'].font = Font(bold=True)
            data_sheet['B6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['C6'].value = 'Average humidity (%)'
            data_sheet['C6'].font = Font(bold=True)
            data_sheet['C6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['D6'].value = 'Average precipitation (mm)'
            data_sheet['D6'].font = Font(bold=True)
            data_sheet['D6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['E6'].value = 'Average wind speed (km/h)'
            data_sheet['E6'].font = Font(bold=True)
            data_sheet['E6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            day_df = filtered_df[filtered_df['Date'] == date].sort_values(by='Time')
            
            data_sheet[f'A{j + 6}'] = date
            data_sheet[f'A{j + 6}'].font = Font(bold=True)
            data_sheet[f'A{j + 6}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            temperature_avg = day_df['Temperature_C'].mean()
            data_sheet[f'B{j + 6}'] = temperature_avg
            data_sheet[f'B{j + 6}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            humididty_avg = day_df['Humidity_pct'].mean()
            data_sheet[f'C{j + 6}'] = humididty_avg
            data_sheet[f'C{j + 6}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            precipitation_avg = day_df['Precipitation_mm'].mean()
            data_sheet[f'D{j + 6}'] = precipitation_avg
            data_sheet[f'D{j + 6}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            wind_speed_avg = day_df['Wind_Speed_kmh'].mean()
            data_sheet[f'E{j + 6}'] = wind_speed_avg
            data_sheet[f'E{j + 6}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            j += 1
            
            wb.save(f'Analyzed/{city}/{months[i].capitalize()}/{months[i].capitalize()}.xlsx')
            
            fig, (ax1, ax2, ax3, ax4) = plt.subplots(4, 1, figsize=(12, 12), sharex=True)
        
            ax1.plot(day_df['Time'], day_df['Temperature_C'], color='blue', marker='o')
            ax1.set_ylabel('Temperature (°C)')
            ax1.grid(True)
            ax1.set_title(f'Weather in {city} for {date}')
            
            ax2.plot(day_df['Time'], day_df['Humidity_pct'], color='green', marker='s')
            ax2.set_ylabel('Humidity (%)')
            ax2.grid(True)
            
            ax3.plot(day_df['Time'], day_df['Precipitation_mm'], color='red', marker='^')
            ax3.set_ylabel('Precipitation (mm)')
            ax3.grid(True)
            
            ax4.plot(day_df['Time'], day_df['Wind_Speed_kmh'], color='yellow', marker='D')
            ax4.set_ylabel('Wind Speed (km/h)')
            ax4.set_xlabel('Time')
            ax4.grid(True)
            
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(f'Analyzed/{city}/{months[i].capitalize()}/{date}.png')
            plt.close()
            
root_folder = "Analyzed"
if os.path.exists(root_folder):
    remove_empty_folders(root_folder)